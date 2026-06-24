### FINAL DYNAMIC MODEL - V9: Total Profit and Output charts
### FINAL DYNAMIC MODEL - V9: Updated charts (Total Profit + Output)
### V7: CORRECTED - Fixes entry timing and promotion accounting bugs from v4/v5
### Entry now contributes to production immediately, promotions properly removed from stock
### V2 calibration (industry-realistic parameters): ÃŽÂ³_AP=1.8%, A/P=3.5
### Stock-flow dynamics with technology shock and comprehensive Excel output

import numpy as np
from scipy.optimize import minimize
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

class DynamicLawFirmModel:
    """
    Dynamic model with stock-flow adjustment for law firms.
    
    Stocks (predetermined):
    - P[t]: Partners (evolves via promotions)
    - A[t]: Associates (evolves via entry and attrition)
    
    Flows (optimized):
    - E[t], R[t], K[t]: Chosen optimally each period
    - entry[t]: New associate hiring
    
    Scenarios:
    A) No ladder: Firm optimizes entry freely ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¾ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ P declines
    B) With ladder: Firm maintains career ladder ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¾ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ P stable
    """
    
    def __init__(self, T=20, shock_period=5, discount_rate=0.05):
        self.T = T
        self.shock_period = shock_period
        self.beta = 1 / (1 + discount_rate)
        
        # Production parameters (raw share weights, no adjustment factors)
        self.alpha_P = 0.35
        self.alpha_K = 0.15
        self.alpha_A = 0.55
        self.alpha_E = 0.25
        
        self.rho_1 = -0.6
        self.Gamma = 1713.0  # Calibrated to AmLaw 100 2024: ~$7M revenue/equity partner
        
        # Wages
        self.w_P = 650.0
        self.w_A = 235.0
        self.w_E = 56.0
        self.r_K = 50.0
        
        # Lateral market premium: tau_bar as fraction of w_P
        # Applied to decentralized scenario post-shock (firms must poach at premium)
        # tau_bar=0.0 = baseline (no premium); 0.10=10%, 0.20=20%, 0.30=30% over w_P
        self.tau_bar = 0.0

        # Technology regimes
        self.sigma_2_baseline = 0.5
        self.rho_2_baseline = (self.sigma_2_baseline - 1) / self.sigma_2_baseline
        self.w_R_baseline = 2.00
        
        self.sigma_2_frontier = 1.5
        self.rho_2_frontier = (self.sigma_2_frontier - 1) / self.sigma_2_frontier
        self.w_R_frontier = 0.50
        
        # Career ladder
        self.gamma_AP = 0.018
        self.delta_P = 0.063
        self.delta_A = 0.16
        
        # Demand
        self.eta = 1.5
        self.A_demand = 9535.05
        
        print("="*80)
        print("DYNAMIC MODEL: LAW FIRM")
        print("="*80)
        print(f"\nTime: T = {T} periods")
        print(f"Shock at: t = {shock_period}")
        print(f"Discount rate: {discount_rate*100:.1f}%")
        print(f"\nInitial: P = 100, A = 350")
        print(f"\nShock: ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚ÂÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€šÃ‚Â ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¾Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¡ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã‚Â¦ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¡: {self.sigma_2_baseline} ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¾ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ {self.sigma_2_frontier}")
        print(f"       w_R: ${self.w_R_baseline} ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Â ÃƒÂ¢Ã¢â€šÂ¬Ã¢â€žÂ¢ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã†â€™ÃƒÂ¢Ã¢â€šÂ¬Ã…Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â ÃƒÆ’Ã†â€™Ãƒâ€ Ã¢â‚¬â„¢ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¡ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¬ÃƒÆ’Ã†â€™Ãƒâ€šÃ‚Â¢ÃƒÆ’Ã‚Â¢ÃƒÂ¢Ã¢â‚¬Å¡Ã‚Â¬Ãƒâ€¦Ã‚Â¾ÃƒÆ’Ã¢â‚¬Å¡Ãƒâ€šÃ‚Â¢ ${self.w_R_frontier}")
    
    def get_regime(self, t):
        if t < self.shock_period:
            return self.rho_2_baseline, self.w_R_baseline
        else:
            return self.rho_2_frontier, self.w_R_frontier
    
    def production(self, P, K, A, E, R, rho_2):
        eps = 1e-10
        P, K, A, E, R = [max(x, eps) for x in [P, K, A, E, R]]
        
        alpha_R = 1 - self.alpha_A - self.alpha_E
        inner = (self.alpha_A * A**rho_2 + 
                self.alpha_E * E**rho_2 + 
                alpha_R * R**rho_2)**(1/rho_2)
        
        alpha_nest = 1 - self.alpha_P - self.alpha_K
        Y = (self.alpha_P * P**self.rho_1 + 
             self.alpha_K * K**self.rho_1 + 
             alpha_nest * inner**self.rho_1)**(1/self.rho_1)
        
        return self.Gamma * Y
    
    def inverse_demand(self, Y):
        return self.A_demand * max(Y, 1e-10)**(-1/self.eta)
    
    def revenue(self, P, K, A, E, R, rho_2):
        Y = self.production(P, K, A, E, R, rho_2)
        price = self.inverse_demand(Y)
        return price * Y
    
    def solve_period(self, P_t, A_t, t, enforce_ladder):
        rho_2, w_R = self.get_regime(t)
        
        # CRITICAL FIX: Before shock, ALWAYS enforce ladder (both scenarios identical)
        if t < self.shock_period:
            enforce_ladder = True
        
        # Effective partner wage: higher in decentralized (no-ladder) post-shock
        # because firms must poach externally at a premium tau_bar over base wage
        if not enforce_ladder and t >= self.shock_period:
            w_P_eff = self.w_P * (1 + self.tau_bar)
        else:
            w_P_eff = self.w_P
        
        if enforce_ladder:
            P_next = (1 - self.delta_P) * P_t + self.gamma_AP * A_t
            A_next_required = (self.delta_P / self.gamma_AP) * P_next
            entry = A_next_required - (1 - self.delta_A) * A_t
            entry = max(entry, 0)
            
            def objective(x):
                K, E, R = x
                if any(xi <= 0 for xi in x): return 1e15
                
                
                A_effective = A_t + entry
                rev = self.revenue(P_t, K, A_effective, E, R, rho_2)
                costs = (self.w_P*P_t + self.w_A*(A_t+entry) + self.w_E*E + 
                        w_R*R + self.r_K*K)
                return -(rev - costs)
                return -(rev - costs)
            
            res = minimize(objective, [500, 300, 50000], 
                          bounds=[(1e-10, None), (1e-10, None), (1e-10, None)],
                          method='L-BFGS-B')
            K, E, R = res.x
            profit_total = -res.fun
            
        else:
            def objective(x):
                K, E, R, entry = x
                if K <= 0 or E <= 0 or R <= 0 or entry < 0: return 1e15
                
                # FIX: Entry contributes to production in period t
                A_effective = A_t + entry
                rev = self.revenue(P_t, K, A_effective, E, R, rho_2)
                costs = (w_P_eff*P_t + self.w_A*(A_t+entry) + self.w_E*E + 
                        w_R*R + self.r_K*K)
                return -(rev - costs)
                return -(rev - costs)
            
            res = minimize(objective, [500, 300, 50000, 50], 
                          bounds=[(1e-10, None), (1e-10, None), (1e-10, None), (0, None)],
                          method='L-BFGS-B')
            K, E, R, entry = res.x
            profit_total = -res.fun
        
        return K, E, R, entry, profit_total, profit_total / P_t
    
    def simulate(self, enforce_ladder=False, P_0=100.0, A_0=350.0):
        results = []
        P_t, A_t = P_0, A_0
        
        for t in range(self.T):
            K_t, E_t, R_t, entry_t, profit_total, profit_pp = self.solve_period(
                P_t, A_t, t, enforce_ladder
            )
            
            rho_2, w_R = self.get_regime(t)
            A_effective_t = A_t + entry_t
            Y_t = self.production(P_t, K_t, A_effective_t, E_t, R_t, rho_2)
            price_t = self.inverse_demand(Y_t)
            
            gap_t = self.gamma_AP * A_t - self.delta_P * P_t
            
            results.append({
                't': t,
                'P': P_t,
                'A_stock': A_t,
                'A_effective': A_t + entry_t,
                'E': E_t,
                'R': R_t,
                'K': K_t,
                'entry': entry_t,
                'Y': Y_t,
                'price': price_t,
                'profit_total': profit_total,
                'profit_per_partner': profit_pp,
                'gap': gap_t,
                'regime': 'baseline' if t < self.shock_period else 'frontier'
            })
            
            P_t = (1 - self.delta_P) * P_t + self.gamma_AP * A_t
            # FIX: Remove promotions from associate stock
            A_t = (1 - self.delta_A) * A_t - self.gamma_AP * A_t + entry_t
        
        return pd.DataFrame(results)
    
    def calculate_npv(self, df):
        """NPV on profit_per_partner (legacy, kept for compatibility)."""
        npv = 0
        for t in range(len(df)):
            npv += (self.beta ** t) * df.loc[t, 'profit_per_partner']
        return npv

    def calculate_npv_total(self, df):
        """NPV on profit_total (primary metric for paper results)."""
        npv = 0
        for t in range(len(df)):
            npv += (self.beta ** t) * df.loc[t, 'profit_total']
        return npv

    def run_analysis(self):
        print("\n" + "="*80)
        print("RUNNING SIMULATIONS")
        print("="*80)
        
        print("\nScenario A: NO CAREER LADDER")
        df_no = self.simulate(enforce_ladder=False)
        
        print("Scenario B: WITH CAREER LADDER")
        df_yes = self.simulate(enforce_ladder=True)
        
        npv_no  = self.calculate_npv(df_no)
        npv_yes = self.calculate_npv(df_yes)

        npv_no_total  = self.calculate_npv_total(df_no)
        npv_yes_total = self.calculate_npv_total(df_yes)

        print("\nNPV (profit_total, primary metric):")
        print(f"  No ladder:   ${npv_no_total:>12,.0f}")
        print(f"  With ladder: ${npv_yes_total:>12,.0f}")
        print(f"  Gap:         ${npv_yes_total - npv_no_total:>12,.0f}")

        # Find horizon at which cumulative NPV gap (profit_total) turns positive
        npv_pos_horizon = None
        for n in range(self.shock_period + 1, len(df_no) + 1):
            gap_n = sum(self.beta**t * (df_yes.loc[t,'profit_total'] - df_no.loc[t,'profit_total'])
                        for t in range(n))
            if gap_n > 0:
                npv_pos_horizon = n
                break
        if npv_pos_horizon:
            print(f"  NPV gap turns positive at planning horizon n={npv_pos_horizon}")
        else:
            print(f"  NPV gap does not turn positive within {len(df_no)}-period horizon")

        print("\nNPV (profit_per_partner, secondary metric):")
        print(f"  No ladder:   ${npv_no:>12,.0f}")
        print(f"  With ladder: ${npv_yes:>12,.0f}")
        
        return df_no, df_yes, npv_no, npv_yes
    
    def plot_results(self, df_no, df_yes):
        fig, axes = plt.subplots(3, 2, figsize=(14, 10))
        fig.suptitle('Dynamic Transition: Law Firm Model', fontsize=16, fontweight='bold')
        
        # Partners
        axes[0, 0].plot(df_no['t'], df_no['P'], 'r-', linewidth=2, label='No Ladder')
        axes[0, 0].plot(df_yes['t'], df_yes['P'], 'b-', linewidth=2, label='With Ladder')
        axes[0, 0].axvline(self.shock_period, color='gray', linestyle='--', alpha=0.5)
        axes[0, 0].set_ylabel('Partners (P)', fontsize=11)
        axes[0, 0].set_title('Partners Evolution', fontweight='bold')
        axes[0, 0].legend()
        axes[0, 0].grid(True, alpha=0.3)
        
        # Associates
        axes[0, 1].plot(df_no['t'], df_no['A_effective'], 'r-', linewidth=2, label='No Ladder')
        axes[0, 1].plot(df_yes['t'], df_yes['A_effective'], 'b-', linewidth=2, label='With Ladder')
        axes[0, 1].axvline(self.shock_period, color='gray', linestyle='--', alpha=0.5)
        axes[0, 1].set_ylabel('Associates (A)', fontsize=11)
        axes[0, 1].set_title('Associates Evolution', fontweight='bold')
        axes[0, 1].legend()
        axes[0, 1].grid(True, alpha=0.3)
        
        # Paralegals
        axes[1, 0].plot(df_no['t'], df_no['E'], 'r-', linewidth=2, label='No Ladder')
        axes[1, 0].plot(df_yes['t'], df_yes['E'], 'b-', linewidth=2, label='With Ladder')
        axes[1, 0].axvline(self.shock_period, color='gray', linestyle='--', alpha=0.5)
        axes[1, 0].set_ylabel('Paralegals (E)', fontsize=11)
        axes[1, 0].set_title('Paralegals Evolution', fontweight='bold')
        axes[1, 0].legend()
        axes[1, 0].grid(True, alpha=0.3)
        
        # Robots (log scale)
        axes[1, 1].plot(df_no['t'], df_no['R'], 'r-', linewidth=2, label='No Ladder')
        axes[1, 1].plot(df_yes['t'], df_yes['R'], 'b-', linewidth=2, label='With Ladder')
        axes[1, 1].axvline(self.shock_period, color='gray', linestyle='--', alpha=0.5)
        axes[1, 1].set_ylabel('Robots (R, log scale)', fontsize=11)
        axes[1, 1].set_yscale('log')
        axes[1, 1].set_title('Automation Evolution', fontweight='bold')
        axes[1, 1].legend()
        axes[1, 1].grid(True, alpha=0.3, which='both')
        
        # Profit per partner
        axes[2, 0].plot(df_no['t'], df_no['profit_total'], 'r-', linewidth=2, label='No Ladder')
        axes[2, 0].plot(df_yes['t'], df_yes['profit_total'], 'b-', linewidth=2, label='With Ladder')
        axes[2, 0].axvline(self.shock_period, color='gray', linestyle='--', alpha=0.5)
        axes[2, 0].set_xlabel('Period', fontsize=11)
        axes[2, 0].set_ylabel('Total Profit ($k)', fontsize=11)
        axes[2, 0].set_title('Total Profit Evolution', fontweight='bold')
        axes[2, 0].legend()
        axes[2, 0].grid(True, alpha=0.3)
        
        # Career ladder gap
        axes[2, 1].plot(df_no['t'], df_no['Y']/1000, 'r-', linewidth=2, label='No Ladder')
        axes[2, 1].plot(df_yes['t'], df_yes['Y']/1000, 'b-', linewidth=2, label='With Ladder')
        axes[2, 1].axvline(self.shock_period, color='gray', linestyle='--', alpha=0.5)
        
        axes[2, 1].set_xlabel('Period', fontsize=11)
        axes[2, 1].set_ylabel('Output Y (thousands)', fontsize=11)
        axes[2, 1].set_title('Output Evolution', fontweight='bold')
        axes[2, 1].legend()
        axes[2, 1].grid(True, alpha=0.3)
        
        plt.tight_layout()
        return fig
    
    def export_to_excel(self, df_no, df_yes, filename):
        """Export comprehensive results to Excel with multiple sheets"""
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Sheet 1: Summary comparison
        ws_summary = wb.create_sheet("Summary")
        
        summary_data = [
            ["LAW FIRMS - DYNAMIC MODEL RESULTS (V4)", "", "", ""],
            ["", "", "", ""],
            ["Calibration: ÃƒÆ’Ã…Â½Ãƒâ€šÃ‚Â³_AP=1.8%, ÃƒÆ’Ã…Â½Ãƒâ€šÃ‚Â´_P=6.3%, ÃƒÆ’Ã…Â½Ãƒâ€šÃ‚Â´_A=16%, A/P=3.5", "", "", ""],
            ["", "", "", ""],
            ["PARAMETER", "VALUE", "", ""],
            ["Promotion rate (ÃƒÆ’Ã…Â½Ãƒâ€šÃ‚Â³_AP)", f"{self.gamma_AP:.4f} ({self.gamma_AP*100:.2f}%)", "", ""],
            ["Partner attrition (ÃƒÆ’Ã…Â½Ãƒâ€šÃ‚Â´_P)", f"{self.delta_P:.4f} ({self.delta_P*100:.2f}%)", "", ""],
            ["Associate attrition (ÃƒÆ’Ã…Â½Ãƒâ€šÃ‚Â´_A)", f"{self.delta_A:.4f} ({self.delta_A*100:.2f}%)", "", ""],
            ["A/P ratio (steady state)", f"{self.delta_P/self.gamma_AP:.2f}", "", ""],
            ["", "", "", ""],
            ["CROSSOVER ANALYSIS", "", "", ""],
            ["Period", "No Ladder Profit", "With Ladder Profit", "Gap"],
        ]
        
        # Find crossover
        crossover = None
        for t in range(20):
            gap = df_yes.loc[t, 'profit_total'] - df_no.loc[t, 'profit_total']
            summary_data.append([t, f"${df_no.loc[t, 'profit_total']:,.0f}", 
                                f"${df_yes.loc[t, 'profit_total']:,.0f}", f"${gap:,.0f}"])
            if t >= self.shock_period and gap > 0 and crossover is None:
                crossover = t
        
        for row in summary_data:
            ws_summary.append(row)
        
        # Format summary sheet
        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary['A5'].font = Font(bold=True)
        ws_summary['A11'].font = Font(bold=True)
        
        # Sheet 2: No Ladder - Full results
        ws_no = wb.create_sheet("No_Ladder")
        ws_no.append(["SCENARIO: NO CAREER LADDER CONSTRAINT"])
        ws_no.append([])
        
        # Column headers
        headers = ['Period', 'Partners', 'Associates', 'Capital', 'Support Staff', 
                  'Automation', 'Entry', 'Output', 'Price', 'Revenue', 
                  'Total Profit', 'Profit/Partner', 'Gap']
        ws_no.append(headers)
        
        for t in range(len(df_no)):
            row = df_no.loc[t]
            ws_no.append([
                t,
                f"{row['P']:.2f}",
                f"{row['A_stock']:.2f}",
                f"{row['K']:.2f}",
                f"{row['E']:.2f}",
                f"{row['R']:.2f}",
                f"{row['entry']:.2f}",
                f"{row['Y']:.2f}",
                f"{row['price']:.2f}",
                f"{row['price'] * row['Y']:.2f}",
                f"{row['profit_total']:.2f}",
                f"{row['profit_per_partner']:.2f}",
                f"{row['gap']:.4f}",
            ])
        
        # Format headers
        for cell in ws_no[3]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Sheet 3: With Ladder - Full results
        ws_yes = wb.create_sheet("With_Ladder")
        ws_yes.append(["SCENARIO: WITH CAREER LADDER CONSTRAINT"])
        ws_yes.append([])
        ws_yes.append(headers)
        
        for t in range(len(df_yes)):
            row = df_yes.loc[t]
            ws_yes.append([
                t,
                f"{row['P']:.2f}",
                f"{row['A_stock']:.2f}",
                f"{row['K']:.2f}",
                f"{row['E']:.2f}",
                f"{row['R']:.2f}",
                f"{row['entry']:.2f}",
                f"{row['Y']:.2f}",
                f"{row['price']:.2f}",
                f"{row['price'] * row['Y']:.2f}",
                f"{row['profit_total']:.2f}",
                f"{row['profit_per_partner']:.2f}",
                f"{row['gap']:.4f}",
            ])
        
        # Format headers
        for cell in ws_yes[3]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Sheet 4: Comparison
        ws_comp = wb.create_sheet("Comparison")
        ws_comp.append(["PERIOD-BY-PERIOD COMPARISON"])
        ws_comp.append([])
        
        comp_headers = ['Period', 'P_No', 'P_Yes', 'P_Diff', 'A_No', 'A_Yes', 'A_Diff',
                       'Profit_No', 'Profit_Yes', 'Profit_Diff', 'PP_No', 'PP_Yes', 'PP_Diff']
        ws_comp.append(comp_headers)
        
        for t in range(20):
            ws_comp.append([
                t,
                f"{df_no.loc[t, 'P']:.2f}",
                f"{df_yes.loc[t, 'P']:.2f}",
                f"{df_yes.loc[t, 'P'] - df_no.loc[t, 'P']:.2f}",
                f"{df_no.loc[t, 'A_stock']:.2f}",
                f"{df_yes.loc[t, 'A_stock']:.2f}",
                f"{df_yes.loc[t, 'A_stock'] - df_no.loc[t, 'A_stock']:.2f}",
                f"{df_no.loc[t, 'profit_total']:.2f}",
                f"{df_yes.loc[t, 'profit_total']:.2f}",
                f"{df_yes.loc[t, 'profit_total'] - df_no.loc[t, 'profit_total']:.2f}",
                f"{df_no.loc[t, 'profit_per_partner']:.2f}",
                f"{df_yes.loc[t, 'profit_per_partner']:.2f}",
                f"{df_yes.loc[t, 'profit_per_partner'] - df_no.loc[t, 'profit_per_partner']:.2f}",
            ])
        
        # Format
        for cell in ws_comp[3]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Save
        wb.save(filename)
        print(f"Excel file saved: {filename}")
        if crossover:
            print(f"ÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã¢â‚¬Å“ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ Crossover detected at period t={crossover}")
        else:
            print(f"ÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã¢â‚¬Å“ÃƒÂ¢Ã¢â€šÂ¬Ã¢â‚¬Â No crossover within 20 periods")


if __name__ == "__main__":
    print("\n")
    
    model = DynamicLawFirmModel(T=60, shock_period=5, discount_rate=0.05)
    df_no, df_yes, npv_no, npv_yes = model.run_analysis()
    
    print("\n" + "="*80)
    print("RESULTS SUMMARY")
    print("="*80)
    
    print("\nNPV (Profit per Partner):")
    print(f"  No ladder:    ${npv_no:,.0f}")
    print(f"  With ladder:  ${npv_yes:,.0f}")
    print(f"  Difference:   ${npv_no - npv_yes:,.0f}")
    
    print("\nKey Statistics (t=19):")
    print(f"\n  NO LADDER:")
    print(f"    Partners: {df_no.loc[19, 'P']:.1f} (from 100)")
    print(f"    Associates: {df_no.loc[19, 'A_effective']:.1f} (from 350)")
    print(f"    Profit/Partner: ${df_no.loc[19, 'profit_per_partner']:,.0f}")
    print(f"    Career gap: {df_no.loc[19, 'gap']:.2f}")
    
    print(f"\n  WITH LADDER:")
    print(f"    Partners: {df_yes.loc[19, 'P']:.1f}")
    print(f"    Associates: {df_yes.loc[19, 'A_effective']:.1f}")
    print(f"    Profit/Partner: ${df_yes.loc[19, 'profit_per_partner']:,.0f}")
    print(f"    Career gap: {df_yes.loc[19, 'gap']:.2f}")
    
    # Check crossover
    crossover = None
    for t in range(20):
        if t >= model.shock_period and df_yes.loc[t, 'profit_total'] > df_no.loc[t, 'profit_total']:
            crossover = t
            break
    
    if crossover:
        print(f"\nÃƒÆ’Ã‚Â¢Ãƒâ€¦Ã¢â‚¬Å“ÃƒÂ¢Ã¢â€šÂ¬Ã…â€œ CROSSOVER at t={crossover}")
        print(f"  At crossover: P_no={df_no.loc[crossover, 'P']:.1f}, P_yes={df_yes.loc[crossover, 'P']:.1f}")
    
    print("\n" + "="*80)
    print("GENERATING OUTPUTS")
    print("="*80)
    
    # Charts
    fig = model.plot_results(df_no, df_yes)
    output_file = '/mnt/user-data/outputs/Law_v4_Charts.png'
    fig.savefig(output_file, dpi=300, bbox_inches='tight')
    print(f"\nChart saved: {output_file}")
    
    # Excel comprehensive output
    excel_file = '/mnt/user-data/outputs/Law_v4_Complete_Results.xlsx'
    model.export_to_excel(df_no, df_yes, excel_file)
    
    print("\n" + "="*80)
